
pip install openpyxl

# path = "/content/KPI_FOLDER/agents1.xlsx"
# path = "/content/KPI_FOLDER/agents1_offline.xlsx"
# path = "/content/KPI_FOLDER/workstations.xlsx"

# import openpyxl module
import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
c1 = sheet.cell(row=1, column=1)

# writing values to cells
c1.value = "computer1"
c2 = sheet.cell(row=2, column=1)
c2.value = "computer2"
c3 = sheet.cell(row=3, column=1)
c3.value = "World111"
wb.save("/content/KPI_FOLDER/workstations.xlsx")
######################################################

wb2 = openpyxl.Workbook()
sheet2 = wb2.active
c11 = sheet2.cell(row=1, column=1)
# writing values to cells
c11.value = "computer1"
# B2 means column = 2 & row = 2.
c22 = sheet2.cell(row=2, column=1)
c22.value = "computer2"
c223 = sheet2.cell(row=3, column=1)
c223.value = "computer23"
wb2.save("/content/KPI_FOLDER/agents1.xlsx")
#########################################################
wb3 = openpyxl.Workbook()
sheet3= wb3.active
ce1 = sheet.cell(row=1, column=1)

# writing values to cells
ce1.value = "computer1 of line"
ce2 = sheet3.cell(row=2, column=1)
ce2.value = "computer2"

wb.save("/content/KPI_FOLDER/agents1_offline.xlsx")
######################################################

# Give the location of the file
path1 = "/content/KPI_FOLDER/workstations.xlsx"
path2 = "/content/KPI_FOLDER/agents1.xlsx"

# To open the workbook
# workbook object is created
wb_obj1 = openpyxl.load_workbook(path1)
wb_obj2 = openpyxl.load_workbook(path2)

# Get workbook active sheet object
# from the active attribute
sheet_obj1 = wb_obj1.active
sheet_obj2 = wb_obj2.active

cell_obj1 = sheet_obj1.cell(row=2, column=1)
cell_obj2 = sheet_obj2.cell(row=2, column=1)

# Give the location of the file
wb_objr1 = openpyxl.load_workbook(path1)
wb_objr2 = openpyxl.load_workbook(path2)


sheet_objr1 = wb_objr1.active
sheet_objr2 = wb_objr2.active

column1 = sheet_objr1.max_row
column2 = sheet_objr2.max_row

print("Total Columns from sheet1:", column1)
print("Total Columns from sheet2:", column2)

print("\nValue of last column1")
for i in range(1, column1+1 ):
	cell_obj1 = sheet_objr1.cell(row=i, column=1)
print(cell_obj1.value)

print("\nValue of last column2")
for j in range(1, column2+1 ):
	cell_obj2 = sheet_obj2.cell(row=i, column=1)
print(cell_obj2.value, end=" ")

# A_infiniti='
# print(A_infiniti)
# cell_obj_11 = sheet_objr1['A1':'B3']
# print (type(cell_obj_11))
# for cell1,cell2 in (cell_obj_11):
#     print(cell1.value)
# print('----------------------------')
# cell_obj_22 = sheet_objr2['A1':'B3']

# for cell11,cell22 in (cell_obj_22):
#     print(cell11.value)

import openpyxl

def read_files(all_computers_file, antivirus_file):

  all_computers_wb = openpyxl.load_workbook(all_computers_file)
  antivirus_wb = openpyxl.load_workbook(antivirus_file)

  # Get the first sheet (assuming computer names are in the first sheet)
  all_computers_sheet = all_computers_wb.active
  antivirus_sheet = antivirus_wb.active

  # Print the first column of all_computers_sheet
  print("First column of all_computers.xlsx:")
  for row in all_computers_sheet.iter_cols(min_col=1, max_col=1):
    for cell in row:
      print(cell.value)

  # Print the first column of antivirus_sheet
  print("\nFirst column of antivirus.xlsx:")
  for row in antivirus_sheet.iter_cols(min_col=1, max_col=1):
    for cell in row:
      print(cell.value)

####--------------------------------------------------------------------------------

def compare_computers(all_computers_file, antivirus_file):

  # Load the workbooks
  all_computers_wb = openpyxl.load_workbook(all_computers_file)
  antivirus_wb = openpyxl.load_workbook(antivirus_file)

  # Get the first sheet (assuming computer names are in the first sheet)
  all_computers_sheet = all_computers_wb.active
  antivirus_sheet = antivirus_wb.active

  # Convert names to lowercase for case-insensitive comparison (optional)
  all_computer_names = []
  for cell in all_computers_sheet.iter_rows(min_col=1, max_col=1):
    # Access cell values directly in the loop
    all_computer_names.append(cell[0].value)


  antivirus_names = []
  for cell in antivirus_sheet.iter_rows(min_col=1, max_col=1):
    antivirus_names.append(cell[0].value)

  # Use set operations to find unique values in the first file not present in the second
  unique_computers = set(all_computer_names) - set(antivirus_names)

  return list(unique_computers)

####--------------------------------------------------------------------------------

def cleanup(all_computers_file, antivirus_file):

  # Load the workbooks
  all_computers_wb_cleanup = openpyxl.load_workbook(all_computers_file)
  antivirus_wb_cleanup = openpyxl.load_workbook(antivirus_file)

  # Get the first sheet (assuming computer names are in the first sheet)
  all_computers_sheet_cleanup = all_computers_wb_cleanup.active
  antivirus_sheet_cleanup = antivirus_wb_cleanup.active

  # Convert names to lowercase for case-insensitive comparison (optional)
  all_computer_names_cleanup = []
  for cell in all_computers_sheet_cleanup.iter_rows(min_col=1, max_col=1):
    # Access cell values directly in the loop
    all_computer_names_cleanup.append(cell[0].value)


  antivirus_names_cleanup = []
  for cell in antivirus_sheet_cleanup.iter_rows(min_col=1, max_col=1):
    antivirus_names_cleanup.append(cell[0].value)

  # Use set operations to find unique values in the first file not present in the second
  cleanup_comp = set(antivirus_names_cleanup) - set(all_computer_names_cleanup)

  return list(cleanup_comp)

#-----------------------------------------------------------------------------


def write_files(list_without_agent, list_tocleanup):

  list_without_agent = list_without_agent
  list_tocleanup = list_tocleanup


  new_wb1 = openpyxl.Workbook()
  new_sheet1 = new_wb1.active
  # Write computer data to rows (assuming one computer per row)
  for row_num1, computer1 in enumerate(list_without_agent, start=1):
    new_sheet1.cell(row=row_num1, column=1).value = computer1
  # Save the workbook as a new xlsx file
  new_wb1.save("/content/KPI_FOLDER/withoutagents.xlsx")



  new_wb2 = openpyxl.Workbook()
  new_sheet2 = new_wb2.active
  # Write computer data to rows (assuming one computer per row)
  for row_num2, computer2 in enumerate(list_tocleanup, start=1):
    new_sheet2.cell(row=row_num2, column=1).value = computer2
  # Save the workbook as a new xlsx file
  new_wb2.save("/content/KPI_FOLDER/cleanup.xlsx")


#-----------------------------------------------------------------------

# Example usage
all_computers_file = "/content/KPI_FOLDER/workstations.xlsx"
antivirus_file = "/content/KPI_FOLDER/agents1.xlsx"

read_files(all_computers_file, antivirus_file)
computers_without_av = compare_computers(all_computers_file, antivirus_file)
computers_needtocleanup = cleanup(all_computers_file, antivirus_file)
write_files(computers_without_av,computers_needtocleanup)
# Print the list of computers without AV
print("Computers without agents ")
for computer in computers_without_av:
  print(computer)


# Print the list oF computers need cleanup
print("list oF computers need to cleanup:")
for computer in computers_needtocleanup:
  print(computer)
#================================================================
import csv

# Paths to save the CSV files
path1 = "C:/content/KPI_FOLDER/workstations.csv"
path2 = "C:/content/KPI_FOLDER/agents1.csv"
path3 = "C:/content/KPI_FOLDER/agents1_offline.csv"

# Data to write into CSV
data1 = [
    ["computer1"],
    ["computer2"],
    ["World111"]
]

data2 = [
    ["computer1"],
    ["computer2"],
    ["computer23"]
]

data3 = [
    ["computer1 of line"],
    ["computer2"]
]

# Writing data1 to workstations.csv
with open(path1, mode='w', newline='') as file:
    writer = csv.writer(file)
    writer.writerows(data1)

# Writing data2 to agents1.csv
with open(path2, mode='w', newline='') as file:
    writer = csv.writer(file)
    writer.writerows(data2)

# Writing data3 to agents1_offline.csv
with open(path3, mode='w', newline='') as file:
    writer = csv.writer(file)
    writer.writerows(data3)

print("CSV files created successfully.")

#+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++

def read_files(all_computers_file, antivirus_file):
    # Reading the CSV files
    with open(all_computers_file, mode='r') as file:
        reader = csv.reader(file)
        all_computers_data = list(reader)

    with open(antivirus_file, mode='r') as file:
        reader = csv.reader(file)
        antivirus_data = list(reader)

    # Print the first column of all_computers.csv
    print("First column of all_computers.csv:")
    for row in all_computers_data:
        print(row[0].strip())  # Trim spaces

    # Print the first column of antivirus.csv
    print("\nFirst column of antivirus.csv:")
    for row in antivirus_data:
        print(row[0].strip())  # Trim spaces

####--------------------------------------------------------------------------------

def compare_computers(all_computers_file, antivirus_file):
    # Reading the CSV files
    with open(all_computers_file, mode='r') as file:
        reader = csv.reader(file)
        all_computers_data = [row[0].strip() for row in list(reader)]  # Trim spaces

    with open(antivirus_file, mode='r') as file:
        reader = csv.reader(file)
        antivirus_data = [row[0].strip() for row in list(reader)]  # Trim spaces

    # Use set operations to find unique values in the first file not present in the second
    unique_computers = set(all_computers_data) - set(antivirus_data)

    return list(unique_computers)

####--------------------------------------------------------------------------------

def cleanup(all_computers_file, antivirus_file):
    # Reading the CSV files
    with open(all_computers_file, mode='r') as file:
        reader = csv.reader(file)
        all_computers_data = [row[0].strip() for row in list(reader)]  # Trim spaces

    with open(antivirus_file, mode='r') as file:
        reader = csv.reader(file)
        antivirus_data = [row[0].strip() for row in list(reader)]  # Trim spaces

    # Use set operations to find unique values in the second file not present in the first
    cleanup_comp = set(antivirus_data) - set(all_computers_data)

    return list(cleanup_comp)

#-----------------------------------------------------------------------------


def write_files(list_without_agent, list_tocleanup):
    # Paths to save the CSV files
    path_without_agent = "C:/content/KPI_FOLDER/withoutagents.csv"
    path_cleanup = "C:/content/KPI_FOLDER/cleanup.csv"

    # Writing list_without_agent to withoutagents.csv
    with open(path_without_agent, mode='w', newline='') as file:
        writer = csv.writer(file)
        for computer in list_without_agent:
            writer.writerow([computer])

    # Writing list_tocleanup to cleanup.csv
    with open(path_cleanup, mode='w', newline='') as file:
        writer = csv.writer(file)
        for computer in list_tocleanup:
            writer.writerow([computer])

    print("CSV files created successfully.")

#-----------------------------------------------------------------------

# Example usage
all_computers_file = "C:/content/KPI_FOLDER/workstations.csv"
antivirus_file = "C:/content/KPI_FOLDER/agents1.csv"

read_files(all_computers_file, antivirus_file)
computers_without_av = compare_computers(all_computers_file, antivirus_file)
computers_needtocleanup = cleanup(all_computers_file, antivirus_file)
write_files(computers_without_av, computers_needtocleanup)

# Print the list of computers without AV
print("Computers without agents:")
for computer in computers_without_av:
    print(computer)

# Print the list of computers that need cleanup
print("List of computers that need to cleanup:")
for computer in computers_needtocleanup:
    print(computer)
#--------------------------------------
import csv

# Input and output file paths
input_file = 'computers.csv'
output_file = 'computers_v1.csv'

# The OS to filter by
filter_os = 'kali'  # Adjust this to the desired OS

# Open the input file in read mode and output file in write mode
with open(input_file, mode='r', newline='') as infile, open(output_file, mode='w', newline='') as outfile:
    reader = csv.DictReader(infile)
    writer = csv.DictWriter(outfile, fieldnames=reader.fieldnames)
    
    # Write the header to the output file
    writer.writeheader()

    # Iterate through the rows and write only those that match the filter
    for row in reader:
        if filter_os.lower() in row['operatingsystem'].lower():  # Case-insensitive match
            writer.writerow(row)

print(f"Filtered data saved to {output_file}")

####-----------asdfasdfasdfasdf------------------------------
#
#
#6
import csv

def filter_and_extract(input_file, output_file):
    """Filters a CSV file based on OS and extracts the 'name' column.

    Args:
        input_file: The path to the input CSV file.
        output_file: The path to the output CSV file.
    """

    with open(input_file, 'r', newline='') as infile, open(output_file, 'w', newline='') as outfile:
        reader = csv.DictReader(infile)
        writer = csv.writer(outfile)
        writer.writerow(['name'])  # Write the header

        for row in reader:
            if row['operatingsystem'] in ['kali', 'windows']:
                writer.writerow([row['name']])

# Example usage:
input_file = 'input.csv'
output_file = 'output.csv'
filter_and_extract(input_file, output_file)


